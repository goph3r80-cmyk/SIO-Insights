"""Generic CSV adapter matching templates/transactions_template.csv.

Column aliases are accepted (case-insensitive): e.g. "trade date"/"date",
"symbol"/"ticker", "amount"/"gross_amount", "commission"/"fees".
"""
from __future__ import annotations

import csv
from pathlib import Path

from ..core import parse_number
from .base import ImportAdapter, ParsedRow

ALIASES = {
    "date": {"date", "trade_date", "trade date", "transaction date", "settled"},
    "type": {"type", "action", "transaction type", "activity"},
    "ticker": {"ticker", "symbol", "code", "counter"},
    "exchange": {"exchange", "market", "mic"},
    "name": {"name", "company", "description", "security"},
    "quantity": {"quantity", "qty", "shares", "units"},
    "price": {"price", "unit price", "price per share", "avg price"},
    "gross_amount": {"gross_amount", "amount", "gross", "value", "total"},
    "fees": {"fees", "fee", "commission", "charges", "brokerage"},
    "tax": {"tax", "withholding", "wht", "withholding tax"},
    "currency": {"currency", "ccy", "cur"},
    "account_broker": {"account_broker", "broker", "platform"},
    "account_name": {"account_name", "account"},
    "account_reference": {"account_reference", "account number", "account no"},
    "note": {"note", "notes", "remarks", "memo"},
    "lot_id": {"lot_id", "lot"},
    "split_ratio": {"split_ratio", "ratio"},
    "external_id": {"external_id", "reference", "ref", "order id", "contract no"},
    "settle_date": {"settle_date", "settlement date"},
}
NUMERIC = {"quantity", "price", "gross_amount", "fees", "tax", "split_ratio"}


def _header_map(fieldnames: list[str]) -> dict[str, str]:
    out = {}
    for col in fieldnames or []:
        key = col.strip().lower().replace("-", " ").replace("_", " ")
        for canon, names in ALIASES.items():
            if key in names or key.replace(" ", "_") in names:
                out[col] = canon
                break
    return out


class GenericCsvAdapter(ImportAdapter):
    name = "generic_csv"

    def sniff(self, path: Path) -> bool:
        if path.suffix.lower() != ".csv":
            return False
        with open(path, newline="", encoding="utf-8-sig") as f:
            header = f.readline().lower()
        return "type" in header and ("date" in header or "trade" in header)

    def parse(self, path: Path) -> list[ParsedRow]:
        rows: list[ParsedRow] = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            hmap = _header_map(reader.fieldnames)
            for i, raw in enumerate(reader, start=2):
                if not any((v or "").strip() for v in raw.values()):
                    continue
                row = ParsedRow(row_number=i, raw=dict(raw))
                if not hmap:
                    row.errors.append("no recognisable columns in header")
                    rows.append(row)
                    continue
                norm: dict = {}
                for col, canon in hmap.items():
                    value = (raw.get(col) or "").strip()
                    if not value:
                        continue
                    if canon in NUMERIC:
                        try:
                            norm[canon] = parse_number(value)
                        except ValueError:
                            row.errors.append(f"'{canon}' is not a number: {value!r}")
                    else:
                        norm[canon] = value
                if norm.get("ticker"):
                    norm["ticker"] = norm["ticker"].upper()
                if norm.get("exchange"):
                    norm["exchange"] = norm["exchange"].upper()
                if norm.get("currency"):
                    norm["currency"] = norm["currency"].upper()
                row.norm = norm
                rows.append(row)
        return rows


class XlsxAdapter(ImportAdapter):
    """XLSX with the same columns as the generic CSV template (first sheet)."""
    name = "generic_xlsx"

    def sniff(self, path: Path) -> bool:
        return path.suffix.lower() in (".xlsx", ".xlsm")

    def parse(self, path: Path) -> list[ParsedRow]:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = [str(c or "") for c in next(it, [])]
        hmap = _header_map(header)
        rows: list[ParsedRow] = []
        for i, values in enumerate(it, start=2):
            raw = {header[j]: values[j] if j < len(values) else None
                   for j in range(len(header))}
            if not any(v not in (None, "") for v in raw.values()):
                continue
            row = ParsedRow(row_number=i, raw={k: str(v) for k, v in raw.items()})
            norm: dict = {}
            for col, canon in hmap.items():
                value = raw.get(col)
                if value in (None, ""):
                    continue
                if canon in NUMERIC:
                    try:
                        norm[canon] = parse_number(value)
                    except ValueError:
                        row.errors.append(f"'{canon}' is not a number: {value!r}")
                elif canon == "date" or canon == "settle_date":
                    norm[canon] = str(value)[:10]
                else:
                    norm[canon] = str(value).strip()
            for k in ("ticker", "exchange", "currency"):
                if norm.get(k):
                    norm[k] = norm[k].upper()
            row.norm = norm
            rows.append(row)
        wb.close()
        return rows


ADAPTERS = {a.name: a for a in (GenericCsvAdapter(), XlsxAdapter())}


def pick_adapter(path: Path, name: str | None = None) -> ImportAdapter:
    if name:
        if name not in ADAPTERS:
            raise ValueError(f"unknown adapter '{name}'; available: {sorted(ADAPTERS)}")
        return ADAPTERS[name]
    for adapter in ADAPTERS.values():
        if adapter.sniff(path):
            return adapter
    if path.suffix.lower() == ".pdf":
        raise ValueError(
            "PDF statements need a broker-specific adapter. Export CSV/XLSX from "
            "your broker, or transcribe rows into the generic CSV template "
            "(templates/transactions_template.csv).")
    raise ValueError(f"no adapter recognises {path.name}; "
                     f"use --adapter with one of {sorted(ADAPTERS)}")
